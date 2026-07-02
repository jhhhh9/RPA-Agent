from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from .workflow import render_template

CORE_RE = re.compile(r"(AB|AC|ER|R)\d+", re.IGNORECASE)


def run_data_pipeline(definition: dict[str, Any], base_row: dict[str, Any]) -> list[dict[str, Any]]:
    rows, _trace = run_data_pipeline_with_trace(definition, base_row)
    return rows


def run_data_pipeline_with_trace(definition: dict[str, Any], base_row: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Run local data preparation and return execution rows plus per-step trace."""
    steps = definition.get("data_pipeline") or []
    if not steps:
        return [base_row], []

    context: dict[str, Any] = {**base_row}
    last_rows: list[dict[str, Any]] | None = None
    trace: list[dict[str, Any]] = []
    for step in steps:
        if not isinstance(step, dict):
            continue
        step_type = str(step.get("type") or "").strip()
        step_id = str(step.get("step_id") or step_type or "step").strip()
        params = render_params(step.get("params") or {}, context)
        input_count = pipeline_input_count(step_type, params, context)
        output_name = ""
        output_value: Any = None
        if step_type == "read_table":
            rows = read_table(Path(str(params.get("path") or "")).expanduser())
            output_name = str(params.get("output") or step.get("step_id") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "group_rows":
            rows = group_rows(dataset(context, params.get("source")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "filter_completed":
            rows = filter_completed(dataset(context, params.get("source")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "derive_regex_list":
            rows = derive_regex_list(dataset(context, params.get("source")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "derive_sku_candidates":
            rows = derive_sku_candidates(dataset(context, params.get("source")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "list_files":
            files = list_files(params)
            output_name = str(params.get("output") or step.get("step_id") or "files")
            context[output_name] = files
            output_value = files
        elif step_type == "find_file":
            path = find_file(fileset(context, params.get("files")), params)
            output_name = str(params.get("output_field") or step.get("step_id") or "file_path")
            context[output_name] = path
            output_value = path
            if last_rows is None:
                last_rows = [{output_name: path}]
            else:
                last_rows = [{**row, output_name: path} for row in last_rows]
                context[str(params.get("output") or params.get("source") or "rows")] = last_rows
        elif step_type == "match_files":
            rows = match_files(dataset(context, params.get("source")), fileset(context, params.get("files")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "validate_required_fields":
            rows = validate_required_fields(dataset(context, params.get("source")), params, context)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        elif step_type == "set_field":
            rows = set_field(dataset(context, params.get("source")), params)
            output_name = str(params.get("output") or params.get("source") or "rows")
            context[output_name] = rows
            output_value = rows
            last_rows = rows
        else:
            raise ValueError(f"unsupported data_pipeline step type: {step_type}")
        trace.append(pipeline_trace_step(step_id, step_type, params, input_count, output_name, output_value))
    return last_rows or [], trace



def pipeline_input_count(step_type: str, params: dict[str, Any], context: dict[str, Any]) -> int:
    if step_type == "read_table":
        return 1 if str(params.get("path") or "").strip() else 0
    if step_type == "list_files":
        root = Path(str(params.get("root") or "")).expanduser()
        return 1 if root.exists() else 0
    if step_type == "find_file":
        return len(fileset(context, params.get("files")))
    if step_type == "match_files":
        return len(dataset(context, params.get("source")))
    if step_type in {"group_rows", "filter_completed", "derive_regex_list", "derive_sku_candidates", "validate_required_fields", "set_field"}:
        return len(dataset(context, params.get("source")))
    return 0


def pipeline_trace_step(
    step_id: str,
    step_type: str,
    params: dict[str, Any],
    input_count: int,
    output_name: str,
    output_value: Any,
) -> dict[str, Any]:
    trace: dict[str, Any] = {
        "step_id": step_id,
        "type": step_type,
        "input_count": input_count,
        "output_count": value_count(output_value),
        "output": output_name,
        "sample": sample_value(output_value),
    }
    if step_type == "match_files":
        trace["match_mode"] = str(params.get("mode") or "name_contains_any")
        trace["values_field"] = str(params.get("values_field") or "")
        trace["output_field"] = str(params.get("output_field") or "")
        trace["pattern"] = str(params.get("pattern") or "")
        trace["target"] = str(params.get("target") or "name")
    if step_type == "list_files":
        trace["root"] = str(params.get("root") or "")
    if step_type == "find_file":
        trace["files"] = str(params.get("files") or "")
        trace["match_mode"] = str(params.get("mode") or "name_contains")
        trace["keyword"] = str(params.get("keyword") or "")
        trace["pattern"] = str(params.get("pattern") or "")
        trace["target"] = str(params.get("target") or "name")
    return trace


def value_count(value: Any) -> int:
    if isinstance(value, list):
        return len(value)
    if value is None:
        return 0
    return 1


def sample_value(value: Any, limit: int = 5) -> list[Any]:
    if isinstance(value, list):
        return value[:limit]
    if value is None:
        return []
    return [value]

def render_params(value: Any, row: dict[str, Any]) -> Any:
    if isinstance(value, dict):
        return {key: render_params(item, row) for key, item in value.items()}
    if isinstance(value, list):
        return [render_params(item, row) for item in value]
    return render_template(value, row, output_dir=str(row.get("output_dir") or ""))


def read_table(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"任务表不存在: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return [normalize_row(row) for row in csv.DictReader(fh)]
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        raise ValueError(f"不支持的任务表格式: {suffix}")
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator, [])]
    rows: list[dict[str, Any]] = []
    for raw in iterator:
        item = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
        if any(str(value or "").strip() for value in item.values()):
            rows.append(normalize_row(item))
    return rows


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key).strip(): value for key, value in row.items() if str(key).strip()}


def group_rows(rows: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    by = str(params.get("by") or "").strip()
    if not by:
        raise ValueError("group_rows requires by")
    collect = params.get("collect")
    collect_map = collect if isinstance(collect, dict) else {}
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = str(row.get(by) or "").strip()
        if key:
            groups[key].append(row)

    out: list[dict[str, Any]] = []
    for key, items in groups.items():
        grouped: dict[str, Any] = {by: key, "rows": items}
        for target, source in collect_map.items():
            values: list[Any] = []
            for item in items:
                raw = item.get(str(source))
                if raw is None or str(raw).strip() == "":
                    continue
                if raw not in values:
                    values.append(raw)
            grouped[str(target)] = values
        out.append(grouped)
    return out


def filter_completed(rows: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    values_field = str(params.get("values_field") or "").strip()
    if not values_field:
        return rows
    done = read_line_set(Path(str(params.get("log_path") or "")).expanduser())
    if not done:
        return rows
    out: list[dict[str, Any]] = []
    for row in rows:
        values = normalize_values(row.get(values_field))
        remaining = [value for value in values if str(value) not in done]
        if not remaining:
            continue
        out.append({**row, values_field: remaining})
    return out


def derive_regex_list(rows: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    source_field = str(params.get("source_field") or "").strip()
    target_field = str(params.get("target_field") or "").strip()
    pattern = str(params.get("pattern") or "").strip()
    if not source_field or not target_field or not pattern:
        raise ValueError("derive_regex_list requires source_field, target_field and pattern")
    regex = re.compile(pattern, re.IGNORECASE)
    unique = bool(params.get("unique", True))
    out: list[dict[str, Any]] = []
    for row in rows:
        values: list[str] = []
        for raw in normalize_values(row.get(source_field)):
            text = str(raw)
            if params.get("normalize_before_match"):
                text = text.upper().replace("-", "")
            for match in regex.finditer(text):
                value = match.group(0).upper()
                if params.get("normalize_core"):
                    value = value.replace("-", "")
                if not unique or value not in values:
                    values.append(value)
        out.append({**row, target_field: values})
    return out


def derive_sku_candidates(rows: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    source_field = str(params.get("source_field") or "").strip()
    target_field = str(params.get("target_field") or "").strip()
    if not source_field or not target_field:
        raise ValueError("derive_sku_candidates requires source_field and target_field")
    formats = [str(item) for item in normalize_values(params.get("formats")) if str(item).strip()]
    if not formats:
        formats = ["compact_full", "dash_full"]
    out: list[dict[str, Any]] = []
    for row in rows:
        candidates: list[str] = []
        for raw in normalize_values(row.get(source_field)):
            for value in sku_candidates(str(raw), formats):
                if value and value not in candidates:
                    candidates.append(value)
        out.append({**row, target_field: candidates})
    return out


def sku_candidates(value: str, formats: list[str]) -> list[str]:
    full = str(value or "").strip().upper()
    if not full:
        return []
    match = re.match(r"^([A-Z]+)-?(\d+)(.*)$", full)
    if not match:
        return [full]
    prefix, number, suffix = match.groups()
    suffix = suffix.lstrip("-")
    compact_suffix = f"-{suffix}" if suffix else ""
    dash_suffix = f"-{suffix}" if suffix else ""
    compact_core = f"{prefix}{number}"
    dash_core = f"{prefix}-{number}"
    compact_full = f"{compact_core}{compact_suffix}"
    dash_full = f"{dash_core}{dash_suffix}"
    mapping = {
        "raw": full,
        "compact_full": compact_full,
        "dash_full": dash_full,
        "compact_core": compact_core,
        "dash_core": dash_core,
    }
    out: list[str] = []
    for fmt in formats:
        item = mapping.get(fmt)
        if item and item not in out:
            out.append(item)
    return out


def list_files(params: dict[str, Any]) -> list[dict[str, Any]]:
    root = Path(str(params.get("root") or "")).expanduser()
    if not root.is_dir():
        return []
    iterator = root.rglob("*") if params.get("recursive") else root.iterdir()
    out: list[dict[str, Any]] = []
    for path in iterator:
        if path.is_file():
            resolved = path.resolve()
            out.append({
                "path": str(resolved),
                "name": path.name,
                "stem": path.stem,
                "suffix": path.suffix.lower(),
            })
    return out


def match_files(rows: list[dict[str, Any]], files: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    values_field = str(params.get("values_field") or "").strip()
    output_field = str(params.get("output_field") or "").strip()
    mode = str(params.get("mode") or "name_contains_any").strip()
    if not values_field or not output_field:
        raise ValueError("match_files requires values_field and output_field")
    out: list[dict[str, Any]] = []
    for row in rows:
        values = [str(value).strip() for value in normalize_values(row.get(values_field)) if str(value).strip()]
        matched: list[str] = []
        for file in files:
            if file_matches(file, values, mode, params):
                path = str(file.get("path") or "")
                if path and path not in matched:
                    matched.append(path)
        out.append({**row, output_field: matched})
    return out


def find_file(files: list[dict[str, Any]], params: dict[str, Any]) -> str:
    mode = str(params.get("mode") or "name_contains").strip()
    keyword = str(params.get("keyword") or "").strip()
    pattern = str(params.get("pattern") or "").strip()
    target = str(params.get("target") or "name").strip()
    case_insensitive = params.get("case_insensitive") is not False
    for file in files:
        text = str(file.get(target) if target in file else file.get("name") or "")
        if single_file_matches(text, mode, keyword, pattern, case_insensitive):
            return str(file.get("path") or "")
    return ""


def single_file_matches(text: str, mode: str, keyword: str, pattern: str, case_insensitive: bool) -> bool:
    haystack = text.lower() if case_insensitive else text
    needle = keyword.lower() if case_insensitive else keyword
    if mode == "regex":
        if not pattern:
            return False
        flags = re.IGNORECASE if case_insensitive else 0
        return bool(re.search(pattern, text, flags))
    if mode == "name_equals":
        return bool(needle and haystack == needle)
    return bool(needle and needle in haystack)


def file_matches(file: dict[str, Any], values: list[str], mode: str, params: dict[str, Any] | None = None) -> bool:
    if not values:
        return False
    params = params or {}
    name = str(file.get("name") or "")
    stem = str(file.get("stem") or "")
    if mode == "core_stem_equals_any":
        core = extract_core(stem)
        return bool(core and core in {extract_core(value) or value.upper() for value in values})
    if mode == "exact_token_any":
        return any(exact_token_match(name, value) for value in values)
    if mode == "regex_any":
        return any(regex_file_match(file, value, params) for value in values)
    return any(value.lower() in name.lower() for value in values)


def regex_file_match(file: dict[str, Any], value: str, params: dict[str, Any]) -> bool:
    target = str(params.get("target") or "name").strip()
    text = str(file.get(target) if target in file else file.get("name") or "")
    pattern = str(params.get("pattern") or "").strip()
    if not pattern:
        return exact_token_match(text, value)
    escaped = re.escape(str(value))
    pattern = pattern.replace("{{value}}", escaped)
    flags = 0 if params.get("case_insensitive") is False else re.IGNORECASE
    return bool(re.search(pattern, text, flags))


def validate_required_fields(rows: list[dict[str, Any]], params: dict[str, Any], context: dict[str, Any]) -> list[dict[str, Any]]:
    required = [str(item) for item in normalize_values(params.get("required_fields"))]
    optional = [str(item) for item in normalize_values(params.get("optional_fields"))]
    log_path = Path(str(params.get("log_path") or "")).expanduser()
    prepare_missing_log(log_path, params)
    out: list[dict[str, Any]] = []
    old_params = context.get("__validate_params")
    context["__validate_params"] = params
    try:
        for row in rows:
            missing_required = [field for field in required if is_empty(row.get(field))]
            missing_optional = [field for field in optional if is_empty(row.get(field))]
            if missing_required or missing_optional:
                write_missing_log(log_path, row, missing_required, missing_optional, context)
            if not missing_required:
                out.append(row)
    finally:
        if old_params is None:
            context.pop("__validate_params", None)
        else:
            context["__validate_params"] = old_params
    return out


def set_field(rows: list[dict[str, Any]], params: dict[str, Any]) -> list[dict[str, Any]]:
    field = str(params.get("field") or "").strip()
    if not field:
        raise ValueError("set_field requires field")
    value = params.get("value")
    return [{**row, field: value} for row in rows]


def dataset(context: dict[str, Any], source: Any) -> list[dict[str, Any]]:
    value = context.get(str(source or ""))
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def fileset(context: dict[str, Any], source: Any) -> list[dict[str, Any]]:
    value = context.get(str(source or ""))
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def normalize_values(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple) or isinstance(value, set):
        return list(value)
    if isinstance(value, str) and ";" in value:
        return [item.strip() for item in value.split(";") if item.strip()]
    return [value]


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) == 0
    return str(value).strip() == ""


def extract_core(value: Any) -> str:
    raw = str(value or "").upper().replace("-", "")
    match = CORE_RE.search(raw)
    return match.group(0) if match else ""


def exact_token_match(file_name: str, target: str) -> bool:
    pattern = re.compile(r"(^|\s|-|_)" + re.escape(str(target)) + r"(\s|-|_|$|\.)", re.IGNORECASE)
    return bool(pattern.search(file_name))


def read_line_set(path: Path) -> set[str]:
    if not str(path) or not path.exists():
        return set()
    return {line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()}


def write_missing_log(
    path: Path,
    row: dict[str, Any],
    missing_required: list[str],
    missing_optional: list[str],
    context: dict[str, Any],
) -> None:
    if not str(path):
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    params = context.get("__validate_params")
    template = ""
    labels: dict[str, str] = {}
    if isinstance(params, dict):
        template = str(params.get("format") or "")
        raw_labels = params.get("field_labels")
        if isinstance(raw_labels, dict):
            labels = {str(key): str(value) for key, value in raw_labels.items()}
    if template:
        line = render_missing_template(template, row, missing_required, missing_optional, labels)
    else:
        label = row.get("SPU ID") or row.get("id") or "-"
        parts = [f"SPU:{label}"]
        sku_list = row.get("sku_list")
        if sku_list:
            parts.append(f"SKU:{format_value(sku_list)}")
        if missing_required:
            parts.append(f"缺失必填:{','.join(label_fields(missing_required, labels))}")
        if missing_optional:
            parts.append(f"缺失可选:{','.join(label_fields(missing_optional, labels))}")
        line = " | ".join(parts)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(line + "\n")


def prepare_missing_log(path: Path, params: dict[str, Any]) -> None:
    if not str(path):
        return
    if str(params.get("write_mode") or "append").strip() != "overwrite":
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("", encoding="utf-8")


def render_missing_template(
    template: str,
    row: dict[str, Any],
    missing_required: list[str],
    missing_optional: list[str],
    labels: dict[str, str],
) -> str:
    values = {
        **{str(key): format_value(value) for key, value in row.items()},
        "missing_required": ",".join(missing_required),
        "missing_optional": ",".join(missing_optional),
        "missing_required_labels": ",".join(label_fields(missing_required, labels)),
        "missing_optional_labels": ",".join(label_fields(missing_optional, labels)),
    }
    out = template
    for key, value in values.items():
        out = out.replace("{" + key + "}", value)
    return out


def label_fields(fields: list[str], labels: dict[str, str]) -> list[str]:
    return [labels.get(field, field) for field in fields]


def format_value(value: Any) -> str:
    if isinstance(value, list):
        return ",".join(str(item) for item in value)
    return str(value)
