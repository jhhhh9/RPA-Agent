from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CORE_RE = re.compile(r'(AB|AC|ER|R)\d+', re.IGNORECASE)


def extract_core_main(text: Any) -> str:
    raw = str(text or '').strip().upper().replace('-', '')
    match = CORE_RE.search(raw)
    return match.group(0) if match else ''


def exact_sku_match(file_name: str, target_sku_full: str) -> bool:
    pattern = re.compile(r'(^|\s|-)' + re.escape(target_sku_full) + r'(\s|-|$|\.)', re.IGNORECASE)
    return bool(pattern.search(file_name))


def load_auto_cert_rows(context: dict[str, Any]) -> list[dict[str, Any]]:
    task_excel = Path(str(context.get('task_excel') or '')).expanduser()
    label_dir = Path(str(context.get('label_dir') or '')).expanduser()
    report_dir = Path(str(context.get('report_dir') or '')).expanduser()
    output_dir = Path(str(context.get('output_dir') or task_excel.parent)).expanduser()
    package_img = normalize_optional_path(context.get('package_img_path'))
    done_log = output_dir / 'done_sku.txt'
    complete_log = output_dir / 'complete_spu.txt'
    miss_log = output_dir / 'miss_resource.txt'
    output_dir.mkdir(parents=True, exist_ok=True)

    spu_group = read_task_excel(task_excel)
    done_set = read_lines(done_log)
    miss_spu_set = {line.split('|')[0].replace('SPU:', '').strip() for line in read_lines(miss_log) if line.strip()}
    rows: list[dict[str, Any]] = []
    for spu_id, sku_list in spu_group.items():
        if spu_id in miss_spu_set:
            continue
        need_run = [sku for sku in sku_list if sku not in done_set]
        if not need_run:
            continue
        unique_core_list = sorted({extract_core_main(sku) for sku in need_run if extract_core_main(sku)})
        label_paths: list[str] = []
        for core in unique_core_list:
            label_paths.extend(find_label_paths(label_dir, core))
        report_paths = collect_report_paths(report_dir, unique_core_list)
        if not label_paths:
            append_line(miss_log, f'SPU:{spu_id} | SKU:{need_run} | 缺失:label')
        if not report_paths:
            append_line(miss_log, f'SPU:{spu_id} | SKU:{need_run} | 缺失:report')
            continue
        rows.append({
            'SPU ID': spu_id,
            'sku_list': need_run,
            'core_sku_list': unique_core_list,
            'label_paths': sorted(set(label_paths)),
            'package_img_path': package_img,
            'report_paths': sorted(set(report_paths)),
            'done_log': str(done_log),
            'complete_spu_log': str(complete_log),
            'miss_resource_log': str(miss_log),
        })
    return rows


def normalize_optional_path(value: Any) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    return str(Path(raw).expanduser().resolve())


def read_task_excel(path: Path) -> dict[str, list[str]]:
    if not path.exists():
        raise FileNotFoundError(f'任务 Excel 不存在: {path}')
    if path.suffix.lower() == '.csv':
        with path.open('r', encoding='utf-8-sig', newline='') as fh:
            reader = csv.DictReader(fh)
            return group_task_rows(reader)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(value or '').strip() for value in rows[0]]
    items = []
    for raw in rows[1:]:
        item = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        items.append(item)
    return group_task_rows(items)


def group_task_rows(rows: Any) -> dict[str, list[str]]:
    out: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        spu = str(row.get('SPU ID') or row.get('spu_id') or '').strip()
        sku = str(row.get('SKU货号') or row.get('sku') or row.get('SKU') or '').strip()
        if spu and sku:
            out[spu].append(sku)
    return dict(out)


def find_label_paths(root: Path, core_sku: str) -> list[str]:
    if not root.is_dir():
        return []
    out: list[str] = []
    for path in root.rglob('*'):
        if path.is_file() and extract_core_main(path.stem) == core_sku:
            out.append(str(path.resolve()))
    return out


def collect_report_paths(root: Path, sku_full_list: list[str]) -> list[str]:
    if not root.is_dir():
        return []
    out: list[str] = []
    for path in root.iterdir():
        if not path.is_file():
            continue
        if any(exact_sku_match(path.name, sku) for sku in sku_full_list):
            out.append(str(path.resolve()))
    return sorted(set(out))


def read_lines(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {line.strip() for line in path.read_text(encoding='utf-8').splitlines() if line.strip()}


def append_line(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('a', encoding='utf-8') as fh:
        fh.write(content + '\n')
